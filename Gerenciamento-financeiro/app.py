#!/usr/bin/env python3
"""
Dashboard Web Interativo - Instituto Areluna
Visualização de Orçamento Empresarial no Navegador
"""

import streamlit as st
import openpyxl
import pandas as pd
from datetime import datetime
import os

# Configuração da página
st.set_page_config(
    page_title="Instituto Areluna - Dashboard Orçamento 2025",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Estilo CSS customizado
st.markdown("""
<style>
    .big-font {
        font-size: 24px !important;
        font-weight: bold;
    }
    .metric-card {
        background-color: #f0f2f6;
        padding: 20px;
        border-radius: 10px;
        margin: 10px 0;
    }
    .positive {
        color: #28a745;
    }
    .negative {
        color: #dc3545;
    }
    .warning {
        color: #ffc107;
    }
</style>
""", unsafe_allow_html=True)


@st.cache_data
def load_workbook(file_path):
    """Carrega o workbook Excel."""
    if not os.path.exists(file_path):
        return None
    return openpyxl.load_workbook(file_path, data_only=True)


@st.cache_data
def read_assumptions(wb):
    """Lê a planilha de premissas."""
    try:
        ws = wb['Assumptions']
        assumptions = {}
        
        # Buscar taxa de crescimento
        for row in range(1, 10):
            cell_label = ws.cell(row, 1).value
            if cell_label and 'Crescimento' in str(cell_label):
                assumptions['growth_rate'] = ws.cell(row, 2).value or 15.0
                break
        
        return assumptions
    except:
        return {'growth_rate': 15.0}


@st.cache_data
def read_dre_data(wb):
    """Lê dados da DRE."""
    try:
        ws = wb['DRE Grencial Instituto ']
    except:
        try:
            ws = wb['DRE Grencial Instituto']
        except:
            return None
    
    # Encontrar linha dos meses
    month_row = 6
    months = ['Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun', 
              'Jul', 'Ago', 'Set', 'Out', 'Nov', 'Dez']
    
    # Ler faturamento (linha 7)
    faturamento = []
    for col in range(3, 15):
        val = ws.cell(7, col).value
        faturamento.append(float(val) if val and isinstance(val, (int, float)) else 0)
    
    # Ler LAIR (linha 48)
    lair = []
    for col in range(3, 15):
        val = ws.cell(48, col).value
        lair.append(float(val) if val and isinstance(val, (int, float)) else 0)
    
    # Ler Lucro Líquido (linha 52)
    lucro_liquido = []
    for col in range(3, 15):
        val = ws.cell(52, col).value
        lucro_liquido.append(float(val) if val and isinstance(val, (int, float)) else 0)
    
    return {
        'months': months,
        'faturamento': faturamento,
        'lair': lair,
        'lucro_liquido': lucro_liquido
    }


@st.cache_data
def read_cashflow_data(wb):
    """Lê dados do fluxo de caixa."""
    try:
        ws = wb['DRE Grencial Instituto ']
    except:
        try:
            ws = wb['DRE Grencial Instituto']
        except:
            return None
    
    months = ['Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun', 
              'Jul', 'Ago', 'Set', 'Out', 'Nov', 'Dez']
    
    # Ler Saldo Acumulado (linha 83)
    saldo_acumulado = []
    for col in range(3, 15):
        val = ws.cell(83, col).value
        saldo_acumulado.append(float(val) if val and isinstance(val, (int, float)) else 100000)
    
    return {
        'months': months,
        'saldo_acumulado': saldo_acumulado
    }


@st.cache_data
def read_scenarios(wb):
    """Lê cenários."""
    try:
        ws = wb['Scenarios']
        scenarios = []
        
        row = 3
        while row < 100:
            cell = ws.cell(row, 1).value
            if cell and 'Cenário:' in str(cell):
                scenario_name = str(cell).replace('Cenário:', '').strip()
                description = ws.cell(row + 1, 1).value
                scenarios.append({
                    'name': scenario_name,
                    'description': description
                })
                row += 2
            row += 1
        
        return scenarios
    except:
        return [
            {'name': 'Base', 'description': 'Cenário base com premissas atuais'},
            {'name': 'Otimista', 'description': 'Cenário otimista: +20% crescimento'},
            {'name': 'Pessimista', 'description': 'Cenário pessimista: +5% crescimento'},
            {'name': 'Conservador', 'description': 'Cenário conservador: +10% crescimento'}
        ]


def format_currency(value):
    """Formata valor como moeda brasileira."""
    if value is None or value == 0:
        return "R$ 0,00"
    return f"R$ {value:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')


def main():
    """Função principal do dashboard."""
    
    # Header
    st.markdown("""
    <h1 style='text-align: center; color: #1F4E78;'>
        📊 Instituto Areluna - Dashboard Orçamento 2025
    </h1>
    """, unsafe_allow_html=True)
    
    st.markdown("---")
    
    # Verificar arquivo
    file_path = "Orçamento Empresarial - Instituto Areluna - modelo-v1.xlsx"
    if not os.path.exists(file_path):
        st.error(f"❌ Arquivo não encontrado: {file_path}")
        st.info("Execute primeiro: `python3 scripts/run_update.py`")
        return
    
    # Carregar workbook
    with st.spinner("Carregando dados..."):
        wb = load_workbook(file_path)
        if not wb:
            st.error("Erro ao carregar workbook")
            return
        
        assumptions = read_assumptions(wb)
        dre_data = read_dre_data(wb)
        cashflow_data = read_cashflow_data(wb)
        scenarios = read_scenarios(wb)
    
    # Sidebar
    st.sidebar.title("🎯 Navegação")
    page = st.sidebar.radio(
        "Selecione a página:",
        ["📈 Visão Geral", "💰 Análise Financeira", "📊 Cenários", "⚙️ Premissas"]
    )
    
    st.sidebar.markdown("---")
    st.sidebar.info(f"""
    **Última atualização:**  
    {datetime.now().strftime('%d/%m/%Y %H:%M')}
    
    **Arquivo:**  
    modelo-v1.xlsx
    """)
    
    # Páginas
    if page == "📈 Visão Geral":
        show_overview(dre_data, cashflow_data, assumptions)
    elif page == "💰 Análise Financeira":
        show_financial_analysis(dre_data, cashflow_data)
    elif page == "📊 Cenários":
        show_scenarios(scenarios, assumptions)
    elif page == "⚙️ Premissas":
        show_assumptions(assumptions)


def show_overview(dre_data, cashflow_data, assumptions):
    """Mostra visão geral."""
    st.header("📈 Visão Geral do Orçamento")
    
    if not dre_data:
        st.warning("Dados da DRE não disponíveis")
        return
    
    # KPIs principais
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        receita_total = sum(dre_data['faturamento'])
        st.metric(
            label="💵 Receita Total 2025",
            value=format_currency(receita_total),
            delta=f"+{assumptions.get('growth_rate', 15)}% vs 2024"
        )
    
    with col2:
        lair_total = sum(dre_data['lair'])
        st.metric(
            label="📊 LAIR Total",
            value=format_currency(lair_total),
            delta=f"{(lair_total/receita_total*100):.1f}% margem" if receita_total > 0 else "0%"
        )
    
    with col3:
        ll_total = sum(dre_data['lucro_liquido'])
        st.metric(
            label="💎 Lucro Líquido",
            value=format_currency(ll_total),
            delta=f"{(ll_total/receita_total*100):.1f}% margem" if receita_total > 0 else "0%"
        )
    
    with col4:
        if cashflow_data:
            saldo_final = cashflow_data['saldo_acumulado'][-1] if cashflow_data['saldo_acumulado'] else 100000
            st.metric(
                label="💰 Saldo Caixa (Dez)",
                value=format_currency(saldo_final),
                delta="Projetado"
            )
    
    st.markdown("---")
    
    # Gráfico de receita mensal
    st.subheader("📊 Receita Mensal 2025")
    
    if dre_data:
        df_receita = pd.DataFrame({
            'Mês': dre_data['months'],
            'Receita': dre_data['faturamento']
        })
        st.bar_chart(df_receita.set_index('Mês')['Receita'])
    
    # Tabela mensal
    st.subheader("📋 Desempenho Mensal")
    
    if dre_data:
        df_mensal = pd.DataFrame({
            'Mês': dre_data['months'],
            'Receita': [format_currency(v) for v in dre_data['faturamento']],
            'LAIR': [format_currency(v) for v in dre_data['lair']],
            'Lucro Líquido': [format_currency(v) for v in dre_data['lucro_liquido']],
        })
        
        if cashflow_data:
            df_mensal['Saldo Caixa'] = [format_currency(v) for v in cashflow_data['saldo_acumulado']]
        
        st.dataframe(df_mensal, use_container_width=True)


def show_financial_analysis(dre_data, cashflow_data):
    """Mostra análise financeira detalhada."""
    st.header("💰 Análise Financeira Detalhada")
    
    if not dre_data:
        st.warning("Dados não disponíveis")
        return
    
    # Tabs
    tab1, tab2, tab3 = st.tabs(["📈 DRE", "💸 Fluxo de Caixa", "📊 Margens"])
    
    with tab1:
        st.subheader("Demonstração do Resultado do Exercício")
        
        df_dre = pd.DataFrame({
            'Mês': dre_data['months'],
            'Faturamento': dre_data['faturamento'],
            'LAIR': dre_data['lair'],
            'Lucro Líquido': dre_data['lucro_liquido']
        })
        
        st.line_chart(df_dre.set_index('Mês'))
        
        # Totais
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("Total Faturamento", format_currency(sum(dre_data['faturamento'])))
        with col2:
            st.metric("Total LAIR", format_currency(sum(dre_data['lair'])))
        with col3:
            st.metric("Total Lucro Líquido", format_currency(sum(dre_data['lucro_liquido'])))
    
    with tab2:
        st.subheader("Fluxo de Caixa Projetado")
        
        if cashflow_data:
            df_cf = pd.DataFrame({
                'Mês': cashflow_data['months'],
                'Saldo Acumulado': cashflow_data['saldo_acumulado']
            })
            
            st.area_chart(df_cf.set_index('Mês'))
            
            # Alertas
            saldo_min = min(cashflow_data['saldo_acumulado'])
            if saldo_min < 50000:
                st.warning(f"⚠️ Alerta: Saldo mínimo de {format_currency(saldo_min)} (abaixo de R$ 50.000)")
            else:
                st.success(f"✅ Saldo sempre acima de R$ 50.000 (mínimo: {format_currency(saldo_min)})")
    
    with tab3:
        st.subheader("Análise de Margens")
        
        receita_total = sum(dre_data['faturamento'])
        lair_total = sum(dre_data['lair'])
        ll_total = sum(dre_data['lucro_liquido'])
        
        if receita_total > 0:
            margem_lair = (lair_total / receita_total) * 100
            margem_liquida = (ll_total / receita_total) * 100
            
            col1, col2 = st.columns(2)
            
            with col1:
                st.metric("Margem LAIR", f"{margem_lair:.2f}%")
                st.progress(max(0, min(100, int(margem_lair))) / 100)
            
            with col2:
                st.metric("Margem Líquida", f"{margem_liquida:.2f}%")
                st.progress(max(0, min(100, int(margem_liquida))) / 100)
        
        # Margens mensais
        margens_mensais = []
        for i, receita in enumerate(dre_data['faturamento']):
            if receita > 0:
                margem = (dre_data['lucro_liquido'][i] / receita) * 100
            else:
                margem = 0
            margens_mensais.append(margem)
        
        df_margens = pd.DataFrame({
            'Mês': dre_data['months'],
            'Margem Líquida (%)': margens_mensais
        })
        
        st.line_chart(df_margens.set_index('Mês'))


def show_scenarios(scenarios, assumptions):
    """Mostra cenários de análise."""
    st.header("📊 Cenários de Análise")
    
    st.info("""
    Os cenários permitem comparar diferentes situações de crescimento, preços e custos.
    Selecione um cenário para ver detalhes.
    """)
    
    # Grid de cenários
    cols = st.columns(2)
    
    for i, scenario in enumerate(scenarios):
        with cols[i % 2]:
            with st.container():
                st.markdown(f"""
                <div class="metric-card">
                    <h3>🎯 {scenario['name']}</h3>
                    <p>{scenario['description']}</p>
                </div>
                """, unsafe_allow_html=True)
    
    st.markdown("---")
    
    # Seletor de cenário
    selected_scenario = st.selectbox(
        "Selecione um cenário para análise detalhada:",
        [s['name'] for s in scenarios]
    )
    
    # Mostrar detalhes (simplificado)
    st.subheader(f"Detalhes: {selected_scenario}")
    
    growth_rates = {
        'Base': 15,
        'Otimista': 20,
        'Pessimista': 5,
        'Conservador': 10
    }
    
    growth = growth_rates.get(selected_scenario, 15)
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.metric("Taxa de Crescimento", f"{growth}%")
    with col2:
        price_adj = {'Otimista': '+10%', 'Pessimista': '-5%'}.get(selected_scenario, '0%')
        st.metric("Ajuste de Preços", price_adj)
    with col3:
        commission = {'Otimista': '12%', 'Pessimista': '8%'}.get(selected_scenario, '10%')
        st.metric("Comissões", commission)


def show_assumptions(assumptions):
    """Mostra premissas."""
    st.header("⚙️ Premissas do Modelo")
    
    st.info("""
    As premissas definem os parâmetros fundamentais do orçamento.
    Para alterá-las, edite a planilha 'Assumptions' no Excel e execute novamente o script.
    """)
    
    # Premissas principais
    col1, col2 = st.columns(2)
    
    with col1:
        st.subheader("📈 Crescimento e Receita")
        st.metric("Taxa de Crescimento 2025", f"{assumptions.get('growth_rate', 15)}%")
        st.text("Sazonalidade: Uniforme (1/12 por mês)")
    
    with col2:
        st.subheader("💰 Prazos e Pagamentos")
        st.text("Recebimento (AR):")
        st.text("• 70% à vista")
        st.text("• 20% em 30 dias")
        st.text("• 10% em 60 dias")
        
        st.text("\nPagamento (AP):")
        st.text("• 80% à vista")
        st.text("• 20% em 30 dias")
    
    st.markdown("---")
    
    # Comissões e Marketing
    col1, col2 = st.columns(2)
    
    with col1:
        st.subheader("💼 Comissões")
        st.metric("Comissão de Vendas", "10%")
        st.metric("Comissão Padrão", "5%")
    
    with col2:
        st.subheader("📢 Marketing")
        st.metric("% da Receita", "5%")
    
    st.markdown("---")
    
    # Mapeamento de produtos
    st.subheader("🔗 Mapeamento Produto → Categoria")
    
    categories = {
        "Odonto e Estética": [
            "Reabilitação Oral", "Harmonização Facial", "Implantologia",
            "Estética Dentária", "Branqueamento Dentário", "Odontopediatria",
            "Ortopedia Facial", "Periodotologia", "Edondontia", "Bruxismo"
        ],
        "Cursos": ["Curso de Capacitação"],
        "Implante Capilar": ["Implante Capilar"]
    }
    
    for category, products in categories.items():
        with st.expander(f"📁 {category} ({len(products)} produtos)"):
            for product in products:
                st.text(f"• {product}")


if __name__ == "__main__":
    main()

