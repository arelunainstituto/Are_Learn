"""
Excel I/O Module for Budget Workbook Operations

Handles reading, writing, and structure preservation for Excel workbooks.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import Optional, Dict, List, Any, Tuple
import os
import shutil
from datetime import datetime
import json


class ExcelWorkbookManager:
    """
    Manages Excel workbook operations with structure preservation.
    """
    
    def __init__(self, file_path: str):
        """
        Initialize the workbook manager.
        
        Args:
            file_path: Path to the Excel workbook
        """
        self.file_path = file_path
        self.workbook = None
        self.backup_path = None
    
    def load(self, data_only: bool = False) -> openpyxl.Workbook:
        """
        Load the workbook.
        
        Args:
            data_only: If True, load cell values instead of formulas
            
        Returns:
            Loaded workbook object
        """
        if not os.path.exists(self.file_path):
            raise FileNotFoundError(f"Workbook not found: {self.file_path}")
        
        self.workbook = openpyxl.load_workbook(self.file_path, data_only=data_only)
        return self.workbook
    
    def save(self, output_path: Optional[str] = None):
        """
        Save the workbook.
        
        Args:
            output_path: Path to save the workbook. If None, overwrites original.
        """
        if self.workbook is None:
            raise ValueError("No workbook loaded. Call load() first.")
        
        save_path = output_path if output_path else self.file_path
        self.workbook.save(save_path)
        print(f"Workbook saved to: {save_path}")
    
    def create_backup(self, backup_dir: str = "backups") -> str:
        """
        Create a backup of the original workbook.
        
        Args:
            backup_dir: Directory to store backups
            
        Returns:
            Path to the backup file
        """
        if not os.path.exists(self.file_path):
            raise FileNotFoundError(f"Source workbook not found: {self.file_path}")
        
        # Create backup directory if it doesn't exist
        os.makedirs(backup_dir, exist_ok=True)
        
        # Generate backup filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        base_name = os.path.splitext(os.path.basename(self.file_path))[0]
        backup_name = f"{base_name}_backup_{timestamp}.xlsx"
        backup_path = os.path.join(backup_dir, backup_name)
        
        # Copy file
        shutil.copy2(self.file_path, backup_path)
        self.backup_path = backup_path
        
        print(f"Backup created: {backup_path}")
        return backup_path
    
    def get_sheet(self, sheet_name: str):
        """
        Get a worksheet by name.
        
        Args:
            sheet_name: Name of the sheet
            
        Returns:
            Worksheet object
        """
        if self.workbook is None:
            raise ValueError("No workbook loaded. Call load() first.")
        
        # Try exact match first
        if sheet_name in self.workbook.sheetnames:
            return self.workbook[sheet_name]
        
        # Try fuzzy match (strip trailing spaces)
        for name in self.workbook.sheetnames:
            if name.strip() == sheet_name.strip():
                return self.workbook[name]
        
        raise ValueError(f"Sheet not found: {sheet_name}. Available: {self.workbook.sheetnames}")
    
    def create_sheet(self, sheet_name: str, index: Optional[int] = None):
        """
        Create a new worksheet.
        
        Args:
            sheet_name: Name for the new sheet
            index: Position to insert the sheet (None = append at end)
            
        Returns:
            New worksheet object
        """
        if self.workbook is None:
            raise ValueError("No workbook loaded. Call load() first.")
        
        if sheet_name in self.workbook.sheetnames:
            print(f"Warning: Sheet '{sheet_name}' already exists. Returning existing sheet.")
            return self.get_sheet(sheet_name)
        
        ws = self.workbook.create_sheet(sheet_name, index)
        return ws
    
    def write_value(self, worksheet, row: int, col: int, value: Any, 
                   number_format: Optional[str] = None):
        """
        Write a value to a cell.
        
        Args:
            worksheet: Target worksheet
            row: Row number (1-based)
            col: Column number (1-based)
            value: Value to write
            number_format: Optional number format string
        """
        cell = worksheet.cell(row=row, column=col)
        cell.value = value
        
        if number_format:
            cell.number_format = number_format
    
    def write_formula(self, worksheet, row: int, col: int, formula: str):
        """
        Write a formula to a cell.
        
        Args:
            worksheet: Target worksheet
            row: Row number (1-based)
            col: Column number (1-based)
            formula: Formula string (without leading =)
        """
        cell = worksheet.cell(row=row, column=col)
        if not formula.startswith('='):
            formula = '=' + formula
        cell.value = formula
    
    def read_range(self, worksheet, start_row: int, start_col: int,
                   end_row: int, end_col: int) -> List[List[Any]]:
        """
        Read a range of cells.
        
        Args:
            worksheet: Source worksheet
            start_row: Starting row (1-based)
            start_col: Starting column (1-based)
            end_row: Ending row (1-based)
            end_col: Ending column (1-based)
            
        Returns:
            2D list of cell values
        """
        data = []
        for row in range(start_row, end_row + 1):
            row_data = []
            for col in range(start_col, end_col + 1):
                cell = worksheet.cell(row=row, column=col)
                row_data.append(cell.value)
            data.append(row_data)
        return data
    
    def write_range(self, worksheet, start_row: int, start_col: int,
                   data: List[List[Any]], preserve_formulas: bool = False):
        """
        Write a range of data to cells.
        
        Args:
            worksheet: Target worksheet
            start_row: Starting row (1-based)
            start_col: Starting column (1-based)
            data: 2D list of values to write
            preserve_formulas: If True, don't overwrite cells with formulas
        """
        for row_idx, row_data in enumerate(data):
            for col_idx, value in enumerate(row_data):
                row = start_row + row_idx
                col = start_col + col_idx
                
                if preserve_formulas:
                    cell = worksheet.cell(row=row, column=col)
                    if isinstance(cell.value, str) and cell.value.startswith('='):
                        continue  # Skip cells with formulas
                
                self.write_value(worksheet, row, col, value)
    
    def apply_header_style(self, worksheet, row: int, start_col: int, end_col: int):
        """
        Apply header styling to a row.
        
        Args:
            worksheet: Target worksheet
            row: Row number
            start_col: First column
            end_col: Last column
        """
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col in range(start_col, end_col + 1):
            cell = worksheet.cell(row=row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
    
    def auto_adjust_column_width(self, worksheet, start_col: int, end_col: int):
        """
        Auto-adjust column widths based on content.
        
        Args:
            worksheet: Target worksheet
            start_col: First column
            end_col: Last column
        """
        for col in range(start_col, end_col + 1):
            column_letter = get_column_letter(col)
            max_length = 0
            
            for row in range(1, worksheet.max_row + 1):
                cell = worksheet.cell(row=row, column=col)
                if cell.value:
                    cell_length = len(str(cell.value))
                    max_length = max(max_length, cell_length)
            
            adjusted_width = min(max_length + 2, 50)  # Cap at 50
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def log_changes(self, log_file: str, changes: Dict[str, Any]):
        """
        Log changes made to the workbook.
        
        Args:
            log_file: Path to log file
            changes: Dictionary of changes to log
        """
        log_entry = {
            "timestamp": datetime.now().isoformat(),
            "workbook": os.path.basename(self.file_path),
            "changes": changes
        }
        
        # Ensure log directory exists
        os.makedirs(os.path.dirname(log_file), exist_ok=True)
        
        # Append to log file
        mode = 'a' if os.path.exists(log_file) else 'w'
        with open(log_file, mode) as f:
            f.write(json.dumps(log_entry, ensure_ascii=False, indent=2) + "\n")


def format_currency_pt(value: float) -> str:
    """
    Format a number as Portuguese currency.
    
    Args:
        value: Numeric value
        
    Returns:
        Formatted string with Portuguese decimal/thousands separators
    """
    if value is None:
        return "0,00"
    
    # Format with 2 decimals
    formatted = f"{value:,.2f}"
    
    # Replace comma and dot for PT format
    formatted = formatted.replace(',', 'TEMP')
    formatted = formatted.replace('.', ',')
    formatted = formatted.replace('TEMP', '.')
    
    return formatted


def parse_currency_pt(value_str: str) -> float:
    """
    Parse a Portuguese-formatted currency string to float.
    
    Args:
        value_str: String with PT format (e.g., "1.234,56")
        
    Returns:
        Float value
    """
    if not value_str or value_str == "":
        return 0.0
    
    # Remove spaces
    cleaned = str(value_str).strip()
    
    # Replace PT separators with EN format
    cleaned = cleaned.replace('.', '')  # Remove thousands separator
    cleaned = cleaned.replace(',', '.')  # Change decimal separator
    
    try:
        return float(cleaned)
    except ValueError:
        return 0.0

