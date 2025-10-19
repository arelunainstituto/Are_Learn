"""
Scenarios Module for Budget Scenario Analysis

Allows creating and comparing different budget scenarios with parameter variations.
"""

from typing import Dict, List, Optional
import copy


class ScenarioManager:
    """
    Manages budget scenarios with parameter deltas.
    """
    
    def __init__(self, assumptions_manager):
        """
        Initialize scenario manager.
        
        Args:
            assumptions_manager: AssumptionsManager instance
        """
        self.base_assumptions = assumptions_manager
        self.scenarios = {}
        self.create_default_scenarios()
    
    def create_default_scenarios(self):
        """Create default scenarios: Base, Optimistic, Pessimistic."""
        # Base scenario (current assumptions)
        self.scenarios['Base'] = {
            'description': 'Cenário base com premissas atuais',
            'deltas': {},
        }
        
        # Optimistic scenario
        self.scenarios['Otimista'] = {
            'description': 'Cenário otimista: +20% crescimento, +10% preços',
            'deltas': {
                'growth_rate_2025': 0.20,  # 20% instead of 15%
                'price_adjustment': 1.10,   # +10% on prices
                'volume_adjustment': 1.05,  # +5% volume
                'commission_pct': 0.12,     # Higher commissions due to better sales
                'marketing_pct': 0.04,      # Lower % as base is higher
            },
        }
        
        # Pessimistic scenario
        self.scenarios['Pessimista'] = {
            'description': 'Cenário pessimista: +5% crescimento, -5% preços',
            'deltas': {
                'growth_rate_2025': 0.05,   # 5% growth
                'price_adjustment': 0.95,    # -5% on prices
                'volume_adjustment': 1.00,   # No volume change
                'commission_pct': 0.08,      # Lower commissions
                'marketing_pct': 0.07,       # Higher % to drive sales
            },
        }
        
        # Conservative scenario
        self.scenarios['Conservador'] = {
            'description': 'Cenário conservador: +10% crescimento, custos mais altos',
            'deltas': {
                'growth_rate_2025': 0.10,
                'price_adjustment': 1.00,
                'volume_adjustment': 1.00,
                'commission_pct': 0.10,
                'marketing_pct': 0.06,
                'csv_adjustment': 1.10,      # +10% on costs
                'fixed_cost_adjustment': 1.05,  # +5% on fixed costs
            },
        }
    
    def add_scenario(self, name: str, description: str, deltas: Dict[str, float]):
        """
        Add a custom scenario.
        
        Args:
            name: Scenario name
            description: Scenario description
            deltas: Dictionary of parameter deltas
        """
        self.scenarios[name] = {
            'description': description,
            'deltas': deltas,
        }
    
    def apply_scenario(self, scenario_name: str, base_values: Dict[str, float]) -> Dict[str, float]:
        """
        Apply scenario deltas to base values.
        
        Args:
            scenario_name: Name of scenario to apply
            base_values: Base values dictionary
            
        Returns:
            Adjusted values dictionary
        """
        if scenario_name not in self.scenarios:
            print(f"Warning: Scenario '{scenario_name}' not found")
            return base_values.copy()
        
        scenario = self.scenarios[scenario_name]
        adjusted_values = base_values.copy()
        
        # Apply deltas
        for key, delta_value in scenario['deltas'].items():
            if key in adjusted_values:
                if 'adjustment' in key:
                    # Multiplicative adjustment
                    adjusted_values[key] = base_values[key] * delta_value
                else:
                    # Direct replacement
                    adjusted_values[key] = delta_value
        
        return adjusted_values
    
    def calculate_scenario_revenue(self, scenario_name: str, base_revenue: float) -> float:
        """
        Calculate revenue for a scenario.
        
        Args:
            scenario_name: Name of scenario
            base_revenue: Base revenue amount
            
        Returns:
            Adjusted revenue
        """
        scenario = self.scenarios.get(scenario_name)
        if not scenario:
            return base_revenue
        
        deltas = scenario['deltas']
        
        # Apply adjustments
        revenue = base_revenue
        
        if 'price_adjustment' in deltas:
            revenue *= deltas['price_adjustment']
        
        if 'volume_adjustment' in deltas:
            revenue *= deltas['volume_adjustment']
        
        return revenue
    
    def create_scenarios_sheet(self, workbook_manager):
        """
        Create the Scenarios sheet in the workbook.
        
        Args:
            workbook_manager: ExcelWorkbookManager instance
        """
        ws = workbook_manager.create_sheet("Scenarios", index=1)
        
        # Title
        from openpyxl.styles import Font
        workbook_manager.write_value(ws, 1, 1, "CENÁRIOS DE ANÁLISE")
        ws.cell(1, 1).font = Font(bold=True, size=14)
        
        row = 3
        
        # Write each scenario
        for scenario_name, scenario_data in self.scenarios.items():
            workbook_manager.write_value(ws, row, 1, f"Cenário: {scenario_name}")
            ws.cell(row, 1).font = Font(bold=True, size=12)
            row += 1
            
            workbook_manager.write_value(ws, row, 1, scenario_data['description'])
            row += 1
            
            # Write deltas
            workbook_manager.write_value(ws, row, 1, "Parâmetro")
            workbook_manager.write_value(ws, row, 2, "Valor")
            row += 1
            
            for param, value in scenario_data['deltas'].items():
                workbook_manager.write_value(ws, row, 1, param)
                
                # Format based on parameter type
                if 'pct' in param or 'rate' in param:
                    workbook_manager.write_value(ws, row, 2, value * 100)
                    ws.cell(row, 2).number_format = '0.00"%"'
                elif 'adjustment' in param:
                    workbook_manager.write_value(ws, row, 2, value)
                    ws.cell(row, 2).number_format = '0.00'
                else:
                    workbook_manager.write_value(ws, row, 2, value)
                
                row += 1
            
            row += 2  # Spacing
        
        # Instructions section
        row += 2
        workbook_manager.write_value(ws, row, 1, "INSTRUÇÕES")
        ws.cell(row, 1).font = Font(bold=True, size=12)
        row += 1
        
        instructions = [
            "1. Modifique os valores dos parâmetros para criar novos cenários",
            "2. Parâmetros com 'adjustment' são multiplicadores (1.0 = sem mudança)",
            "3. Parâmetros com 'rate' ou 'pct' são percentuais",
            "4. Execute o script de atualização para aplicar os cenários",
        ]
        
        for instruction in instructions:
            workbook_manager.write_value(ws, row, 1, instruction)
            row += 1
        
        # Auto-adjust columns
        workbook_manager.auto_adjust_column_width(ws, 1, 2)
        
        print("Scenarios sheet created successfully")
    
    def get_scenario_comparison(self, base_metrics: Dict[str, float]) -> Dict[str, Dict[str, float]]:
        """
        Compare all scenarios against base metrics.
        
        Args:
            base_metrics: Base case metrics
            
        Returns:
            Dictionary with scenario comparisons
        """
        comparison = {}
        
        for scenario_name in self.scenarios.keys():
            comparison[scenario_name] = {}
            
            # Calculate key metrics with scenario adjustments
            scenario = self.scenarios[scenario_name]
            deltas = scenario['deltas']
            
            # Revenue
            revenue = base_metrics.get('revenue', 0)
            if 'growth_rate_2025' in deltas:
                base_growth = self.base_assumptions.assumptions['growth_rate_2025']
                revenue = revenue / (1 + base_growth) * (1 + deltas['growth_rate_2025'])
            
            revenue = self.calculate_scenario_revenue(scenario_name, revenue)
            comparison[scenario_name]['revenue'] = revenue
            
            # Costs
            csv = base_metrics.get('csv', 0)
            if 'csv_adjustment' in deltas:
                csv *= deltas['csv_adjustment']
            comparison[scenario_name]['csv'] = csv
            
            fixed = base_metrics.get('fixed_costs', 0)
            if 'fixed_cost_adjustment' in deltas:
                fixed *= deltas['fixed_cost_adjustment']
            comparison[scenario_name]['fixed_costs'] = fixed
            
            # Variable costs
            variable = base_metrics.get('variable_costs', 0)
            marketing_pct = deltas.get('marketing_pct', 0.05)
            commission_pct = deltas.get('commission_pct', 0.10)
            variable = revenue * (marketing_pct + commission_pct)
            comparison[scenario_name]['variable_costs'] = variable
            
            # Margins
            lb = revenue - csv
            comparison[scenario_name]['lucro_bruto'] = lb
            comparison[scenario_name]['margem_bruta_pct'] = (lb / revenue * 100) if revenue > 0 else 0
            
            lair = lb - fixed - variable
            comparison[scenario_name]['lair'] = lair
            comparison[scenario_name]['margem_lair_pct'] = (lair / revenue * 100) if revenue > 0 else 0
        
        return comparison
    
    def get_scenario_list(self) -> List[str]:
        """Get list of all scenario names."""
        return list(self.scenarios.keys())
    
    def get_scenario_details(self, scenario_name: str) -> Optional[Dict]:
        """
        Get details for a specific scenario.
        
        Args:
            scenario_name: Name of scenario
            
        Returns:
            Scenario dictionary or None
        """
        return self.scenarios.get(scenario_name)

