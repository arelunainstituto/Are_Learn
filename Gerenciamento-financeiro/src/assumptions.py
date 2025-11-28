"""
Assumptions Module for Budget Model

Creates and manages the Assumptions sheet with all key parameters.
"""

from typing import Dict, List, Any
import json


class AssumptionsManager:
    """
    Manages budget assumptions and parameters.
    """
    
    # Default product to category mapping
    DEFAULT_PRODUCT_MAPPING = {
        "Reabilitação Oral": "Odonto e Estética",
        "Harmonização Facial": "Odonto e Estética",
        "Alinhadores Invisiveis": "Odonto e Estética",
        "Laserterapia": "Odonto e Estética",
        "Implantologia": "Odonto e Estética",
        "Estética Dentária": "Odonto e Estética",
        "Branqueamento Dentário": "Odonto e Estética",
        "Odontopediatria": "Odonto e Estética",
        "Ortopedia Facial": "Odonto e Estética",
        "Periodotologia": "Odonto e Estética",
        "Edondontia": "Odonto e Estética",
        "Bruxismo": "Odonto e Estética",
        "Apneia do Sono": "Odonto e Estética",
        "Bichectomia": "Odonto e Estética",
        "Cirurigia Oral": "Odonto e Estética",
        "Higiene Oral": "Odonto e Estética",
        "Ortodontia": "Odonto e Estética",
        "Curso de Capacitação": "Cursos",
        "Implante Capilar": "Implante Capilar",
    }
    
    def __init__(self):
        """Initialize with default assumptions."""
        self.assumptions = {
            "growth_rate_2025": 0.15,
            "monthly_seasonality": [1/12] * 12,  # Uniform distribution by default
            "tax_rates": {
                "sales_tax": 0.0,  # Placeholder
                "income_tax": 0.0,  # Placeholder
            },
            "ar_terms": {
                "days": [0, 30, 60],
                "weights": [0.7, 0.2, 0.1]
            },
            "ap_terms": {
                "days": [0, 30],
                "weights": [0.8, 0.2]
            },
            "commission_defaults": {
                "sales_commission_pct": 0.10,  # 10% default
                "default_pct": 0.05
            },
            "marketing_defaults": {
                "marketing_pct_revenue": 0.05  # 5% of revenue
            },
            "product_to_category_map": self.DEFAULT_PRODUCT_MAPPING.copy()
        }
    
    def set_growth_rate(self, rate: float):
        """Set annual growth rate."""
        self.assumptions["growth_rate_2025"] = rate
    
    def set_seasonality(self, weights: List[float]):
        """
        Set monthly seasonality weights.
        
        Args:
            weights: List of 12 weights (should sum to 1.0)
        """
        if len(weights) != 12:
            raise ValueError("Seasonality must have exactly 12 weights")
        
        total = sum(weights)
        if abs(total - 1.0) > 0.01:
            # Normalize if not summing to 1
            weights = [w / total for w in weights]
        
        self.assumptions["monthly_seasonality"] = weights
    
    def set_tax_rates(self, sales_tax: float, income_tax: float):
        """Set tax rates."""
        self.assumptions["tax_rates"]["sales_tax"] = sales_tax
        self.assumptions["tax_rates"]["income_tax"] = income_tax
    
    def set_ar_terms(self, days: List[int], weights: List[float]):
        """Set accounts receivable terms."""
        if len(days) != len(weights):
            raise ValueError("AR days and weights must have same length")
        
        total = sum(weights)
        if abs(total - 1.0) > 0.01:
            weights = [w / total for w in weights]
        
        self.assumptions["ar_terms"]["days"] = days
        self.assumptions["ar_terms"]["weights"] = weights
    
    def set_ap_terms(self, days: List[int], weights: List[float]):
        """Set accounts payable terms."""
        if len(days) != len(weights):
            raise ValueError("AP days and weights must have same length")
        
        total = sum(weights)
        if abs(total - 1.0) > 0.01:
            weights = [w / total for w in weights]
        
        self.assumptions["ap_terms"]["days"] = days
        self.assumptions["ap_terms"]["weights"] = weights
    
    def add_product_mapping(self, product: str, category: str):
        """Add or update product to category mapping."""
        self.assumptions["product_to_category_map"][product] = category
    
    def get_category_for_product(self, product: str) -> str:
        """Get category for a product."""
        return self.assumptions["product_to_category_map"].get(product, "Odonto e Estética")
    
    def create_assumptions_sheet(self, workbook_manager):
        """
        Create the Assumptions sheet in the workbook.
        
        Args:
            workbook_manager: ExcelWorkbookManager instance
        """
        ws = workbook_manager.create_sheet("Assumptions", index=0)
        
        # Title
        from openpyxl.styles import Font
        workbook_manager.write_value(ws, 1, 1, "PREMISSAS DO ORÇAMENTO 2025")
        ws.cell(1, 1).font = Font(bold=True, size=14)
        
        row = 3
        
        # Growth Rate
        workbook_manager.write_value(ws, row, 1, "Taxa de Crescimento 2025 (%)")
        workbook_manager.write_value(ws, row, 2, self.assumptions["growth_rate_2025"] * 100)
        ws.cell(row, 2).number_format = '0.00'
        row += 2
        
        # Seasonality
        workbook_manager.write_value(ws, row, 1, "Sazonalidade Mensal")
        row += 1
        months = ['Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho',
                  'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro']
        
        for i, (month, weight) in enumerate(zip(months, self.assumptions["monthly_seasonality"])):
            workbook_manager.write_value(ws, row + i, 1, month)
            workbook_manager.write_value(ws, row + i, 2, weight)
            ws.cell(row + i, 2).number_format = '0.0000'
        row += 13
        
        # Tax Rates
        workbook_manager.write_value(ws, row, 1, "Alíquotas de Impostos")
        row += 1
        workbook_manager.write_value(ws, row, 1, "Impostos sobre Vendas (%)")
        workbook_manager.write_value(ws, row, 2, self.assumptions["tax_rates"]["sales_tax"] * 100)
        ws.cell(row, 2).number_format = '0.00'
        row += 1
        workbook_manager.write_value(ws, row, 1, "Imposto de Renda (%)")
        workbook_manager.write_value(ws, row, 2, self.assumptions["tax_rates"]["income_tax"] * 100)
        ws.cell(row, 2).number_format = '0.00'
        row += 2
        
        # AR Terms
        workbook_manager.write_value(ws, row, 1, "Prazos de Recebimento (AR)")
        row += 1
        workbook_manager.write_value(ws, row, 1, "Dias")
        workbook_manager.write_value(ws, row, 2, "Peso")
        row += 1
        for days, weight in zip(self.assumptions["ar_terms"]["days"], 
                                self.assumptions["ar_terms"]["weights"]):
            workbook_manager.write_value(ws, row, 1, days)
            workbook_manager.write_value(ws, row, 2, weight)
            ws.cell(row, 2).number_format = '0.00'
            row += 1
        row += 1
        
        # AP Terms
        workbook_manager.write_value(ws, row, 1, "Prazos de Pagamento (AP)")
        row += 1
        workbook_manager.write_value(ws, row, 1, "Dias")
        workbook_manager.write_value(ws, row, 2, "Peso")
        row += 1
        for days, weight in zip(self.assumptions["ap_terms"]["days"],
                                self.assumptions["ap_terms"]["weights"]):
            workbook_manager.write_value(ws, row, 1, days)
            workbook_manager.write_value(ws, row, 2, weight)
            ws.cell(row, 2).number_format = '0.00'
            row += 1
        row += 1
        
        # Commission Defaults
        workbook_manager.write_value(ws, row, 1, "Comissões Padrão")
        row += 1
        workbook_manager.write_value(ws, row, 1, "Comissão de Vendas (%)")
        workbook_manager.write_value(ws, row, 2, 
                                     self.assumptions["commission_defaults"]["sales_commission_pct"] * 100)
        ws.cell(row, 2).number_format = '0.00'
        row += 1
        workbook_manager.write_value(ws, row, 1, "Comissão Padrão (%)")
        workbook_manager.write_value(ws, row, 2,
                                     self.assumptions["commission_defaults"]["default_pct"] * 100)
        ws.cell(row, 2).number_format = '0.00'
        row += 2
        
        # Marketing Defaults
        workbook_manager.write_value(ws, row, 1, "Marketing")
        row += 1
        workbook_manager.write_value(ws, row, 1, "% da Receita")
        workbook_manager.write_value(ws, row, 2,
                                     self.assumptions["marketing_defaults"]["marketing_pct_revenue"] * 100)
        ws.cell(row, 2).number_format = '0.00'
        row += 2
        
        # Product to Category Mapping
        workbook_manager.write_value(ws, row, 1, "Mapeamento Produto → Categoria")
        row += 1
        workbook_manager.write_value(ws, row, 1, "Produto")
        workbook_manager.write_value(ws, row, 2, "Categoria")
        row += 1
        
        for product, category in sorted(self.assumptions["product_to_category_map"].items()):
            workbook_manager.write_value(ws, row, 1, product)
            workbook_manager.write_value(ws, row, 2, category)
            row += 1
        
        # Auto-adjust column widths
        workbook_manager.auto_adjust_column_width(ws, 1, 2)
        
        print("Assumptions sheet created successfully")
    
    def save_to_json(self, file_path: str):
        """Save assumptions to JSON file."""
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(self.assumptions, f, ensure_ascii=False, indent=2)
    
    def load_from_json(self, file_path: str):
        """Load assumptions from JSON file."""
        with open(file_path, 'r', encoding='utf-8') as f:
            self.assumptions = json.load(f)
    
    def load_from_sheet(self, worksheet):
        """
        Load assumptions from an existing Assumptions sheet.
        
        Args:
            worksheet: The Assumptions worksheet
        """
        # This can be implemented to read back from the sheet
        # For now, we'll use default values
        pass

