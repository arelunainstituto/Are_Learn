"""
Dashboard Module for KPI Visualization

Creates a dashboard sheet with key performance indicators.
"""

from typing import Dict, List, Optional
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class DashboardManager:
    """
    Creates and manages the dashboard sheet with KPIs.
    """
    
    def __init__(self):
        """Initialize dashboard manager."""
        self.kpis = {}
    
    def create_dashboard_sheet(self, workbook_manager, dre_data: Dict, 
                               cashflow_data: Dict, product_data: Optional[Dict] = None):
        """
        Create the dashboard (Painel) sheet.
        
        Args:
            workbook_manager: ExcelWorkbookManager instance
            dre_data: DRE monthly data
            cashflow_data: Cash flow data
            product_data: Optional product-level data
        """
        ws = workbook_manager.create_sheet("Painel", index=0)
        
        # Title
        ws.merge_cells('A1:G1')
        workbook_manager.write_value(ws, 1, 1, "PAINEL DE INDICADORES - INSTITUTO ARELUNA 2025")
        ws.cell(1, 1).font = Font(bold=True, size=16, color="FFFFFF")
        ws.cell(1, 1).fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        ws.cell(1, 1).alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30
        
        row = 3
        
        # Section 1: Financial Summary
        self._create_section_header(workbook_manager, ws, row, 1, "RESUMO FINANCEIRO ANUAL")
        row += 1
        
        # Calculate annual totals
        annual_revenue = sum(dre_data.get('faturamento', [0] * 12))
        annual_lb = sum(dre_data.get('lucro_bruto', [0] * 12))
        annual_lair = sum(dre_data.get('lair', [0] * 12))
        annual_ll = sum(dre_data.get('lucro_liquido', [0] * 12))
        
        kpis_financial = [
            ("Receita Total", annual_revenue, "R$ #,##0.00"),
            ("Lucro Bruto", annual_lb, "R$ #,##0.00"),
            ("LAIR", annual_lair, "R$ #,##0.00"),
            ("Lucro Líquido", annual_ll, "R$ #,##0.00"),
        ]
        
        row = self._write_kpi_section(workbook_manager, ws, row, 1, kpis_financial)
        row += 2
        
        # Section 2: Margins
        self._create_section_header(workbook_manager, ws, row, 1, "MARGENS")
        row += 1
        
        margem_bruta = (annual_lb / annual_revenue * 100) if annual_revenue > 0 else 0
        margem_lair = (annual_lair / annual_revenue * 100) if annual_revenue > 0 else 0
        margem_liquida = (annual_ll / annual_revenue * 100) if annual_revenue > 0 else 0
        
        kpis_margins = [
            ("Margem Bruta", margem_bruta, "0.00%"),
            ("Margem LAIR", margem_lair, "0.00%"),
            ("Margem Líquida", margem_liquida, "0.00%"),
        ]
        
        row = self._write_kpi_section(workbook_manager, ws, row, 1, kpis_margins)
        row += 2
        
        # Section 3: Cash Flow
        self._create_section_header(workbook_manager, ws, row, 1, "FLUXO DE CAIXA")
        row += 1
        
        if cashflow_data:
            saldo_inicial = 100000  # From assumptions
            saldo_final = cashflow_data.get('accumulated_balance', [0] * 12)[-1] if cashflow_data.get('accumulated_balance') else 0
            saldo_min = min(cashflow_data.get('accumulated_balance', [0])) if cashflow_data.get('accumulated_balance') else 0
            geracao_caixa = sum(cashflow_data.get('monthly_balance', [0] * 12))
            
            kpis_cashflow = [
                ("Saldo Inicial", saldo_inicial, "R$ #,##0.00"),
                ("Geração de Caixa (Total)", geracao_caixa, "R$ #,##0.00"),
                ("Saldo Final Projetado", saldo_final, "R$ #,##0.00"),
                ("Saldo Mínimo no Período", saldo_min, "R$ #,##0.00"),
            ]
            
            row = self._write_kpi_section(workbook_manager, ws, row, 1, kpis_cashflow)
        
        row += 2
        
        # Section 4: Monthly Performance Table
        self._create_section_header(workbook_manager, ws, row, 1, "DESEMPENHO MENSAL")
        row += 1
        
        # Headers
        headers = ["Mês", "Receita", "Lucro Bruto", "LAIR", "Lucro Líquido", "Margem LB%", "Saldo Caixa"]
        for col, header in enumerate(headers, 1):
            workbook_manager.write_value(ws, row, col, header)
            ws.cell(row, col).font = Font(bold=True)
            ws.cell(row, col).fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        
        row += 1
        
        # Monthly data
        months = ['Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun', 
                 'Jul', 'Ago', 'Set', 'Out', 'Nov', 'Dez']
        
        monthly_revenue = dre_data.get('faturamento', [0] * 12)
        monthly_lb = dre_data.get('lucro_bruto', [0] * 12)
        monthly_lair = dre_data.get('lair', [0] * 12)
        monthly_ll = dre_data.get('lucro_liquido', [0] * 12)
        
        accumulated_cash = [100000]  # Start with initial balance
        if cashflow_data and cashflow_data.get('accumulated_balance'):
            accumulated_cash = [100000] + cashflow_data['accumulated_balance']
        
        for i, month in enumerate(months):
            workbook_manager.write_value(ws, row, 1, month)
            
            revenue = monthly_revenue[i] if i < len(monthly_revenue) else 0
            lb = monthly_lb[i] if i < len(monthly_lb) else 0
            lair = monthly_lair[i] if i < len(monthly_lair) else 0
            ll = monthly_ll[i] if i < len(monthly_ll) else 0
            margem_lb = (lb / revenue * 100) if revenue > 0 else 0
            cash = accumulated_cash[i + 1] if (i + 1) < len(accumulated_cash) else 0
            
            workbook_manager.write_value(ws, row, 2, revenue)
            ws.cell(row, 2).number_format = 'R$ #,##0.00'
            
            workbook_manager.write_value(ws, row, 3, lb)
            ws.cell(row, 3).number_format = 'R$ #,##0.00'
            
            workbook_manager.write_value(ws, row, 4, lair)
            ws.cell(row, 4).number_format = 'R$ #,##0.00'
            
            workbook_manager.write_value(ws, row, 5, ll)
            ws.cell(row, 5).number_format = 'R$ #,##0.00'
            
            workbook_manager.write_value(ws, row, 6, margem_lb)
            ws.cell(row, 6).number_format = '0.00%'
            
            workbook_manager.write_value(ws, row, 7, cash)
            ws.cell(row, 7).number_format = 'R$ #,##0.00'
            
            # Color code cash balance
            if cash < 50000:
                ws.cell(row, 7).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            elif cash > 150000:
                ws.cell(row, 7).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            
            row += 1
        
        # Total row
        workbook_manager.write_value(ws, row, 1, "TOTAL")
        ws.cell(row, 1).font = Font(bold=True)
        
        for col in range(2, 6):
            formula = f"=SUM({get_column_letter(col)}{row-12}:{get_column_letter(col)}{row-1})"
            workbook_manager.write_formula(ws, row, col, formula)
            ws.cell(row, col).font = Font(bold=True)
            ws.cell(row, col).number_format = 'R$ #,##0.00'
        
        row += 3
        
        # Section 5: Product Performance (if available)
        if product_data:
            self._create_section_header(workbook_manager, ws, row, 1, "TOP PRODUTOS POR RECEITA")
            row += 1
            
            # Sort products by revenue
            products_sorted = sorted(product_data.items(), 
                                   key=lambda x: x[1].get('receita_liquida', 0), 
                                   reverse=True)
            
            # Headers
            headers = ["Produto", "Receita Líquida", "CSV", "Margem Bruta", "MB%"]
            for col, header in enumerate(headers, 1):
                workbook_manager.write_value(ws, row, col, header)
                ws.cell(row, col).font = Font(bold=True)
                ws.cell(row, col).fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            
            row += 1
            
            # Top 10 products
            for product, data in products_sorted[:10]:
                receita = data.get('receita_liquida', 0)
                csv = data.get('csv', 0)
                lb = receita - csv
                mb_pct = (lb / receita * 100) if receita > 0 else 0
                
                workbook_manager.write_value(ws, row, 1, product)
                workbook_manager.write_value(ws, row, 2, receita)
                ws.cell(row, 2).number_format = 'R$ #,##0.00'
                workbook_manager.write_value(ws, row, 3, csv)
                ws.cell(row, 3).number_format = 'R$ #,##0.00'
                workbook_manager.write_value(ws, row, 4, lb)
                ws.cell(row, 4).number_format = 'R$ #,##0.00'
                workbook_manager.write_value(ws, row, 5, mb_pct)
                ws.cell(row, 5).number_format = '0.00%'
                
                row += 1
        
        # Auto-adjust column widths
        workbook_manager.auto_adjust_column_width(ws, 1, 7)
        
        print("Dashboard sheet created successfully")
    
    def _create_section_header(self, workbook_manager, ws, row: int, col: int, title: str):
        """Create a section header."""
        workbook_manager.write_value(ws, row, col, title)
        ws.cell(row, col).font = Font(bold=True, size=12, color="FFFFFF")
        ws.cell(row, col).fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws.cell(row, col).alignment = Alignment(horizontal="left", vertical="center")
    
    def _write_kpi_section(self, workbook_manager, ws, row: int, col: int, 
                          kpis: List[tuple]) -> int:
        """
        Write a section of KPIs.
        
        Args:
            workbook_manager: ExcelWorkbookManager instance
            ws: Worksheet
            row: Starting row
            col: Starting column
            kpis: List of (label, value, format) tuples
            
        Returns:
            Next row number
        """
        for label, value, number_format in kpis:
            workbook_manager.write_value(ws, row, col, label)
            ws.cell(row, col).font = Font(bold=True)
            
            workbook_manager.write_value(ws, row, col + 1, value)
            ws.cell(row, col + 1).number_format = number_format
            
            # Color code based on value
            if isinstance(value, (int, float)):
                if value < 0:
                    ws.cell(row, col + 1).font = Font(color="FF0000")
                elif value > 0:
                    ws.cell(row, col + 1).font = Font(color="006100")
            
            row += 1
        
        return row
    
    def calculate_variance_vs_target(self, actual: Dict[str, float], 
                                    target: Dict[str, float]) -> Dict[str, Dict[str, float]]:
        """
        Calculate variance between actual and target values.
        
        Args:
            actual: Actual values
            target: Target values
            
        Returns:
            Dictionary with variance calculations
        """
        variance = {}
        
        for key in target.keys():
            if key in actual:
                actual_value = actual[key]
                target_value = target[key]
                
                variance[key] = {
                    'actual': actual_value,
                    'target': target_value,
                    'variance': actual_value - target_value,
                    'variance_pct': ((actual_value - target_value) / target_value * 100) 
                                   if target_value != 0 else 0,
                }
        
        return variance

